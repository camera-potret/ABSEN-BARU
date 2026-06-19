import os
import json
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
import qrcode
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.utils import secure_filename
import uuid

app = Flask(__name__, template_folder='app/templates', static_folder='app/static')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///absensi.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

db = SQLAlchemy(app)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ==================== DATABASE MODELS ====================

class Kegiatan(db.Model):
    __tablename__ = 'kegiatan'
    id = db.Column(db.Integer, primary_key=True)
    nama = db.Column(db.String(255), nullable=False)
    deskripsi = db.Column(db.Text)
    jadwal = db.Column(db.DateTime)
    link_id = db.Column(db.String(50), unique=True, nullable=False)
    wa_group_link = db.Column(db.String(255))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    pesertum = db.relationship('Peserta', backref='kegiatan', lazy=True, cascade='all, delete-orphan')
    absens = db.relationship('Absensi', backref='kegiatan', lazy=True, cascade='all, delete-orphan')

class Peserta(db.Model):
    __tablename__ = 'peserta'
    id = db.Column(db.Integer, primary_key=True)
    kegiatan_id = db.Column(db.Integer, db.ForeignKey('kegiatan.id'), nullable=False)
    nama = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    absens = db.relationship('Absensi', backref='peserta', lazy=True, cascade='all, delete-orphan')

class Absensi(db.Model):
    __tablename__ = 'absensi'
    id = db.Column(db.Integer, primary_key=True)
    kegiatan_id = db.Column(db.Integer, db.ForeignKey('kegiatan.id'), nullable=False)
    peserta_id = db.Column(db.Integer, db.ForeignKey('peserta.id'), nullable=False)
    status = db.Column(db.String(20), default='belum_masuk')  # masuk, terlambat, izin, belum_masuk
    alasan_izin = db.Column(db.Text)
    waktu_absensi = db.Column(db.DateTime, default=datetime.utcnow)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ==================== ROUTES ====================

@app.before_request
def create_tables():
    db.create_all()

@app.route('/')
def dashboard():
    return render_template('dashboard.html')

@app.route('/absensi/<link_id>')
def absensi_page(link_id):
    kegiatan = Kegiatan.query.filter_by(link_id=link_id).first()
    if not kegiatan:
        return "Link tidak valid", 404
    return render_template('absensi.html', link_id=link_id)

# ==================== API ENDPOINTS ====================

# Kegiatan
@app.route('/api/kegiatan', methods=['GET'])
def get_kegiatan():
    kegiatan = Kegiatan.query.first()
    if not kegiatan:
        return jsonify(None), 200
    
    return jsonify({
        'id': kegiatan.id,
        'nama': kegiatan.nama,
        'deskripsi': kegiatan.deskripsi,
        'jadwal': kegiatan.jadwal.isoformat() if kegiatan.jadwal else None,
        'link_id': kegiatan.link_id,
        'wa_group_link': kegiatan.wa_group_link,
        'created_at': kegiatan.created_at.isoformat(),
        'updated_at': kegiatan.updated_at.isoformat()
    }), 200

@app.route('/api/kegiatan', methods=['POST'])
def create_or_update_kegiatan():
    data = request.json
    
    kegiatan = Kegiatan.query.first()
    
    if not kegiatan:
        link_id = data.get('link_id', str(uuid.uuid4())[:8])
        kegiatan = Kegiatan(
            nama=data['nama'],
            deskripsi=data.get('deskripsi'),
            jadwal=datetime.fromisoformat(data['jadwal']) if data.get('jadwal') else None,
            link_id=link_id,
            wa_group_link=data.get('wa_group_link')
        )
        db.session.add(kegiatan)
    else:
        old_link_id = kegiatan.link_id
        kegiatan.nama = data['nama']
        kegiatan.deskripsi = data.get('deskripsi')
        kegiatan.jadwal = datetime.fromisoformat(data['jadwal']) if data.get('jadwal') else None
        kegiatan.wa_group_link = data.get('wa_group_link')
        
        if data.get('link_id') and data.get('link_id') != old_link_id:
            kegiatan.link_id = data.get('link_id')
        
        kegiatan.updated_at = datetime.utcnow()
    
    db.session.commit()
    return jsonify({'success': True, 'kegiatan_id': kegiatan.id}), 200

# Peserta
@app.route('/api/peserta', methods=['GET'])
def get_peserta():
    kegiatan = Kegiatan.query.first()
    if not kegiatan:
        return jsonify([]), 200
    
    peserta = Peserta.query.filter_by(kegiatan_id=kegiatan.id).all()
    
    result = []
    for p in peserta:
        absensi = Absensi.query.filter_by(peserta_id=p.id).first()
        status = absensi.status if absensi else 'belum_masuk'
        alasan_izin = absensi.alasan_izin if absensi and absensi.alasan_izin else ''
        
        result.append({
            'id': p.id,
            'nama': p.nama,
            'status': status,
            'alasan_izin': alasan_izin,
            'waktu_absensi': absensi.waktu_absensi.isoformat() if absensi and absensi.waktu_absensi else None
        })
    
    return jsonify(result), 200

@app.route('/api/peserta/import', methods=['POST'])
def import_peserta():
    if 'file' not in request.files:
        return jsonify({'error': 'File tidak ditemukan'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'File tidak dipilih'}), 400
    
    kegiatan = Kegiatan.query.first()
    if not kegiatan:
        return jsonify({'error': 'Kegiatan belum dibuat'}), 400
    
    try:
        workbook = openpyxl.load_workbook(file)
        worksheet = workbook.active
        
        # Clear existing peserta
        Peserta.query.filter_by(kegiatan_id=kegiatan.id).delete()
        Absensi.query.filter_by(kegiatan_id=kegiatan.id).delete()
        db.session.commit()
        
        # Import new peserta
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            nama = row[0].value
            if nama:
                peserta = Peserta(kegiatan_id=kegiatan.id, nama=str(nama).strip())
                db.session.add(peserta)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'Data berhasil diimport'}), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/peserta/search/<link_id>', methods=['GET'])
def search_peserta(link_id):
    query = request.args.get('q', '').strip()
    
    kegiatan = Kegiatan.query.filter_by(link_id=link_id).first()
    if not kegiatan:
        return jsonify({'error': 'Link tidak valid'}), 404
    
    peserta = Peserta.query.filter_by(kegiatan_id=kegiatan.id).filter(
        Peserta.nama.ilike(f'%{query}%')
    ).all()
    
    result = []
    for p in peserta:
        absensi = Absensi.query.filter_by(peserta_id=p.id).first()
        if not absensi or absensi.status == 'belum_masuk':
            result.append({
                'id': p.id,
                'nama': p.nama
            })
    
    return jsonify(result), 200

# Absensi
@app.route('/api/absensi/masuk', methods=['POST'])
def absensi_masuk():
    data = request.json
    peserta_id = data.get('peserta_id')
    link_id = data.get('link_id')
    
    kegiatan = Kegiatan.query.filter_by(link_id=link_id).first()
    if not kegiatan:
        return jsonify({'error': 'Link tidak valid'}), 404
    
    peserta = Peserta.query.get(peserta_id)
    if not peserta or peserta.kegiatan_id != kegiatan.id:
        return jsonify({'error': 'Peserta tidak ditemukan'}), 404
    
    absensi = Absensi.query.filter_by(peserta_id=peserta_id).first()
    
    jadwal = kegiatan.jadwal
    now = datetime.utcnow()
    
    if jadwal and now > jadwal:
        status = 'terlambat'
    else:
        status = 'masuk'
    
    if absensi:
        absensi.status = status
        absensi.waktu_absensi = now
    else:
        absensi = Absensi(
            kegiatan_id=kegiatan.id,
            peserta_id=peserta_id,
            status=status,
            waktu_absensi=now
        )
        db.session.add(absensi)
    
    db.session.commit()
    
    return jsonify({'success': True, 'status': status}), 200

@app.route('/api/absensi/izin', methods=['POST'])
def absensi_izin():
    data = request.json
    peserta_id = data.get('peserta_id')
    link_id = data.get('link_id')
    alasan = data.get('alasan', '')
    
    kegiatan = Kegiatan.query.filter_by(link_id=link_id).first()
    if not kegiatan:
        return jsonify({'error': 'Link tidak valid'}), 404
    
    peserta = Peserta.query.get(peserta_id)
    if not peserta or peserta.kegiatan_id != kegiatan.id:
        return jsonify({'error': 'Peserta tidak ditemukan'}), 404
    
    absensi = Absensi.query.filter_by(peserta_id=peserta_id).first()
    
    if absensi:
        absensi.status = 'izin'
        absensi.alasan_izin = alasan
    else:
        absensi = Absensi(
            kegiatan_id=kegiatan.id,
            peserta_id=peserta_id,
            status='izin',
            alasan_izin=alasan
        )
        db.session.add(absensi)
    
    db.session.commit()
    
    return jsonify({'success': True}), 200

# Dashboard Stats
@app.route('/api/stats', methods=['GET'])
def get_stats():
    kegiatan = Kegiatan.query.first()
    if not kegiatan:
        return jsonify({
            'total_peserta': 0,
            'masuk': 0,
            'terlambat': 0,
            'izin': 0,
            'belum_masuk': 0
        }), 200
    
    total = Peserta.query.filter_by(kegiatan_id=kegiatan.id).count()
    masuk = db.session.query(func.count(Absensi.id)).filter_by(
        kegiatan_id=kegiatan.id, status='masuk'
    ).scalar() or 0
    terlambat = db.session.query(func.count(Absensi.id)).filter_by(
        kegiatan_id=kegiatan.id, status='terlambat'
    ).scalar() or 0
    izin = db.session.query(func.count(Absensi.id)).filter_by(
        kegiatan_id=kegiatan.id, status='izin'
    ).scalar() or 0
    belum_masuk = total - masuk - terlambat - izin
    
    return jsonify({
        'total_peserta': total,
        'masuk': masuk,
        'terlambat': terlambat,
        'izin': izin,
        'belum_masuk': belum_masuk
    }), 200

# QR Code
@app.route('/api/qrcode', methods=['GET'])
def generate_qrcode():
    kegiatan = Kegiatan.query.first()
    if not kegiatan:
        return jsonify({'error': 'Kegiatan belum dibuat'}), 400
    
    base_url = request.host_url.rstrip('/')
    qr_url = f"{base_url}/absensi/{kegiatan.link_id}"
    
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(qr_url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    img_io = BytesIO()
    img.save(img_io, 'PNG')
    img_io.seek(0)
    
    return send_file(img_io, mimetype='image/png', as_attachment=False)

@app.route('/api/qrcode/download', methods=['GET'])
def download_qrcode():
    kegiatan = Kegiatan.query.first()
    if not kegiatan:
        return jsonify({'error': 'Kegiatan belum dibuat'}), 400
    
    base_url = request.host_url.rstrip('/')
    qr_url = f"{base_url}/absensi/{kegiatan.link_id}"
    
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(qr_url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    img_io = BytesIO()
    img.save(img_io, 'PNG')
    img_io.seek(0)
    
    return send_file(
        img_io,
        mimetype='image/png',
        as_attachment=True,
        download_name=f'qrcode_{kegiatan.link_id}.png'
    )

# Export Excel
@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    kegiatan = Kegiatan.query.first()
    if not kegiatan:
        return jsonify({'error': 'Kegiatan belum dibuat'}), 400
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Absensi"
    
    # Header
    headers = ['No.', 'Nama', 'Status', 'Alasan Izin', 'Waktu Absensi']
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    peserta = Peserta.query.filter_by(kegiatan_id=kegiatan.id).all()
    for idx, p in enumerate(peserta, 1):
        absensi = Absensi.query.filter_by(peserta_id=p.id).first()
        status = absensi.status if absensi else 'belum_masuk'
        alasan = absensi.alasan_izin if absensi and absensi.alasan_izin else ''
        waktu = absensi.waktu_absensi.strftime('%Y-%m-%d %H:%M:%S') if absensi and absensi.waktu_absensi else ''
        
        worksheet.cell(row=idx+1, column=1).value = idx
        worksheet.cell(row=idx+1, column=2).value = p.nama
        worksheet.cell(row=idx+1, column=3).value = status
        worksheet.cell(row=idx+1, column=4).value = alasan
        worksheet.cell(row=idx+1, column=5).value = waktu
    
    # Adjust column width
    worksheet.column_dimensions['A'].width = 5
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 30
    worksheet.column_dimensions['E'].width = 20
    
    excel_io = BytesIO()
    workbook.save(excel_io)
    excel_io.seek(0)
    
    return send_file(
        excel_io,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'absensi_{kegiatan.nama}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

# Reset Data
@app.route('/api/reset', methods=['POST'])
def reset_data():
    kegiatan = Kegiatan.query.first()
    if kegiatan:
        Absensi.query.filter_by(kegiatan_id=kegiatan.id).delete()
        Peserta.query.filter_by(kegiatan_id=kegiatan.id).delete()
        db.session.delete(kegiatan)
        db.session.commit()
    
    return jsonify({'success': True}), 200

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
