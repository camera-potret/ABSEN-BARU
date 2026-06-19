#!/usr/bin/env python3
"""
Script untuk membuat file Excel contoh untuk import peserta
Run: python create_sample_excel.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create workbook
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Peserta"

# Add header
header = ["Nama Peserta"]
worksheet.append(header)

# Style header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for cell in worksheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add sample data
sample_data = [
    ["Andi Pratama"],
    ["Budi Santoso"],
    ["Citra Dewi"],
    ["Dodi Hermawan"],
    ["Eka Putri"],
    ["Fajar Irawan"],
    ["Gina Susanti"],
    ["Hendra Wijaya"],
    ["Intan Kusuma"],
    ["Joko Supriyanto"],
]

for row_idx, row_data in enumerate(sample_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Set column width
worksheet.column_dimensions['A'].width = 25

# Save file
output_file = "sample_peserta.xlsx"
workbook.save(output_file)

print(f"✅ File '{output_file}' berhasil dibuat!")
print(f"   Anda bisa menggunakan file ini sebagai contoh untuk import peserta.")
print(f"\n📝 Cara menggunakan:")
print(f"   1. Edit file dan tambahkan nama peserta")
print(f"   2. Buka dashboard di http://localhost:5000")
print(f"   3. Klik 'Import Excel' dan pilih file ini")
